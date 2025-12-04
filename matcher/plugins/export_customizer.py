
from __future__ import annotations
import os
from typing import Dict, Any

import pandas as pd

try:
    from openpyxl.utils.dataframe import dataframe_to_rows
except Exception:
    def dataframe_to_rows(df, index=False, header=True):
        rows = []
        if header:
            rows.append(list(df.columns))
        for _, row in df.iterrows():
            rows.append(list(row.values))
        return rows

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from datetime import datetime

from ..constants import PREFIX_T1, PREFIX_T2, LABEL_GSS, LABEL_OPERA

PLUGIN_NAME = "Export Customizer"
PLUGIN_DESCRIPTION = (
    "Creates Excel export with essential and full sheets, "
    "color-coded by Confidence."
)
PLUGIN_STAGE = "post_match"


def post_match(df: pd.DataFrame, ctx: Dict[str, Any]) -> None:
    export_folder = ctx.get("export_folder") or ""
    base = ctx.get("export_base") or "match_results"
    log = ctx.get("log") or (lambda msg: None)

    if not export_folder or not os.path.isdir(export_folder):
        log("[ExportCustomizer] No valid export folder; skipping.")
        return

    # Debug: log available columns
    t1_name_cols = [c for c in df.columns if PREFIX_T1 in c and ("name" in c.lower() or "Name" in c)]
    if t1_name_cols:
        log(f"[ExportCustomizer] Found T1 name columns: {', '.join(t1_name_cols)}")
    else:
        log(f"[ExportCustomizer] No T1 name columns found. Available T1 columns: {[c for c in df.columns if c.startswith(PREFIX_T1)]}")

    path = os.path.join(export_folder, f"{base}_v7_4_1.xlsx")

    # Helper to find column with case-insensitive matching
    def find_col(possible_names):
        for name in possible_names:
            if name in df.columns:
                return name
        # Try case-insensitive
        df_cols_lower = {c.lower(): c for c in df.columns}
        for name in possible_names:
            if name.lower() in df_cols_lower:
                return df_cols_lower[name.lower()]
        return None

    # Build essential columns list with flexible matching
    essential_cols = []
    
    # T1 (GSS) columns - try common variations
    t1_last = find_col([f"{PREFIX_T1}Last Name", f"{PREFIX_T1}Last name", f"{PREFIX_T1}LastName", f"{PREFIX_T1}last name"])
    if not t1_last:
        # Fallback: find any T1 column with "last" and "name" in it
        for col in df.columns:
            if col.startswith(PREFIX_T1) and "last" in col.lower() and "name" in col.lower():
                t1_last = col
                break
    if t1_last:
        essential_cols.append(t1_last)
    
    t1_first = find_col([f"{PREFIX_T1}First Name", f"{PREFIX_T1}First name", f"{PREFIX_T1}FirstName", f"{PREFIX_T1}first name"])
    if not t1_first:
        # Fallback: find any T1 column with "first" and "name" in it
        for col in df.columns:
            if col.startswith(PREFIX_T1) and "first" in col.lower() and "name" in col.lower():
                t1_first = col
                break
    if t1_first:
        essential_cols.append(t1_first)
    
    t1_arrival = find_col([f"{PREFIX_T1}Arrival Date", f"{PREFIX_T1}Arrival date", f"{PREFIX_T1}arrival date"])
    if t1_arrival:
        essential_cols.append(t1_arrival)
    
    t1_departure = find_col([f"{PREFIX_T1}Departure Date", f"{PREFIX_T1}Departure date", f"{PREFIX_T1}departure date"])
    if t1_departure:
        essential_cols.append(t1_departure)
    
    t1_itr = find_col([f"{PREFIX_T1}Intent to Recommend (Property)", f"{PREFIX_T1}Intent to Recommend (property)"])
    if t1_itr:
        essential_cols.append(t1_itr)
    
    if "ITR_Bucket" in df.columns:
        essential_cols.append("ITR_Bucket")
    
    t1_loyalty = find_col([f"{PREFIX_T1}Loyalty Program Tier", f"{PREFIX_T1}Loyalty Program tier", f"{PREFIX_T1}Loyalty program tier"])
    if t1_loyalty:
        essential_cols.append(t1_loyalty)
    
    t1_comment = find_col([f"{PREFIX_T1}Overall Comment", f"{PREFIX_T1}Overall comment", f"{PREFIX_T1}overall comment"])
    if t1_comment:
        essential_cols.append(t1_comment)
    
    # T2 (Opera) columns
    t2_userid = find_col([f"{PREFIX_T2}USERID", f"{PREFIX_T2}UserID", f"{PREFIX_T2}userid", f"{PREFIX_T2}User ID"])
    if t2_userid:
        essential_cols.append(t2_userid)
    
    t2_last = find_col([f"{PREFIX_T2}LastName", f"{PREFIX_T2}Last Name", f"{PREFIX_T2}last name"])
    if t2_last:
        essential_cols.append(t2_last)
    
    t2_first = find_col([f"{PREFIX_T2}FirstName", f"{PREFIX_T2}First Name", f"{PREFIX_T2}first name"])
    if t2_first:
        essential_cols.append(t2_first)
    
    t2_date = find_col([f"{PREFIX_T2}Date", f"{PREFIX_T2}date"])
    if t2_date:
        essential_cols.append(t2_date)
    
    # Confidence column (but not score columns)
    if "Confidence" in df.columns:
        essential_cols.append("Confidence")
    
    essential_cols = [c for c in essential_cols if c in df.columns]
    essential_df = df[essential_cols].copy()
    
    # Rename columns: T1_ -> GSS, T2_ -> Opera
    rename_map = {}
    for col in essential_df.columns:
        if col.startswith(PREFIX_T1):
            rename_map[col] = col.replace(PREFIX_T1, LABEL_GSS, 1)
        elif col.startswith(PREFIX_T2):
            rename_map[col] = col.replace(PREFIX_T2, LABEL_OPERA, 1)
    essential_df = essential_df.rename(columns=rename_map)

    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Main"

    for r in dataframe_to_rows(essential_df, index=False, header=True):
        ws_main.append(r)

    # Format date columns and auto-size columns
    header_row = 1
    date_cols = []
    for idx, col_name in enumerate(essential_df.columns, start=1):
        col_letter = get_column_letter(idx)
        # Check if column name suggests it's a date
        if "date" in col_name.lower() or "arrival" in col_name.lower() or "departure" in col_name.lower():
            date_cols.append(col_letter)
            # Auto-size date columns
            ws_main.column_dimensions[col_letter].width = 12
    
    # Format date cells
    from openpyxl.styles import numbers
    for col_letter in date_cols:
        for row in ws_main.iter_rows(min_row=2, min_col=ord(col_letter) - ord('A') + 1, max_col=ord(col_letter) - ord('A') + 1):
            for cell in row:
                if cell.value is not None:
                    # Handle different date types
                    if isinstance(cell.value, str):
                        # Try to parse string dates
                        try:
                            # Try common date formats
                            for fmt in ["%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y", "%d.%m.%y"]:
                                try:
                                    dt = datetime.strptime(cell.value, fmt)
                                    cell.value = dt
                                    cell.number_format = numbers.FORMAT_DATE_DDMMYY
                                    break
                                except ValueError:
                                    continue
                            # If string parsing failed, try pandas to_datetime
                            if isinstance(cell.value, str) and not isinstance(cell.value, datetime):
                                try:
                                    dt_val = pd.to_datetime(cell.value, dayfirst=False, errors='coerce')
                                    if pd.notna(dt_val):
                                        cell.value = dt_val.to_pydatetime() if hasattr(dt_val, 'to_pydatetime') else dt_val
                                        cell.number_format = numbers.FORMAT_DATE_DDMMYY
                                except:
                                    pass
                        except:
                            pass
                    elif isinstance(cell.value, (datetime, pd.Timestamp)):
                        # Already a datetime, just format it
                        if isinstance(cell.value, pd.Timestamp):
                            cell.value = cell.value.to_pydatetime()
                        cell.number_format = numbers.FORMAT_DATE_DDMMYY
                    elif isinstance(cell.value, (int, float)):
                        # Might be Excel serial number
                        try:
                            # Excel serial dates start from 1900-01-01
                            if 1 <= cell.value <= 1000000:  # Reasonable date range
                                dt = from_excel(cell.value)
                                cell.value = dt
                                cell.number_format = numbers.FORMAT_DATE_DDMMYY
                        except:
                            pass
                    elif hasattr(cell.value, 'date'):
                        # Date-like object
                        cell.value = cell.value.date() if hasattr(cell.value, 'date') else cell.value
                        cell.number_format = numbers.FORMAT_DATE_DDMMYY
    
    # Auto-size other columns
    for idx, col_name in enumerate(essential_df.columns, start=1):
        col_letter = get_column_letter(idx)
        if col_letter not in date_cols:
            # Estimate width based on column name and some sample data
            max_len = len(str(col_name))
            if len(essential_df) > 0:
                sample_len = max([len(str(val)) for val in essential_df[col_name].head(10).values if val is not None], default=0)
                max_len = max(max_len, sample_len)
            ws_main.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)
    conf_col_idx = None
    for idx, cell in enumerate(ws_main[header_row], start=1):
        if cell.value == "Confidence":
            conf_col_idx = idx
            break

    fill_high = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    fill_med = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )
    fill_low = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    fill_very_low = PatternFill(
        start_color="FF9999", end_color="FF9999", fill_type="solid"
    )
    fill_no_match = PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    )

    if conf_col_idx:
        for row in ws_main.iter_rows(min_row=2):
            c = row[conf_col_idx - 1]
            v = str(c.value or "").strip().lower()
            if v == "high":
                for cell in row:
                    cell.fill = fill_high
            elif v == "medium":
                for cell in row:
                    cell.fill = fill_med
            elif v == "low":
                for cell in row:
                    cell.fill = fill_low
            elif v == "very low":
                for cell in row:
                    cell.fill = fill_very_low
            elif v == "no match":
                for cell in row:
                    cell.fill = fill_no_match

    ws_full = wb.create_sheet("FullData")
    # Remove score columns from full data and rename T1_/T2_ prefixes
    full_df = df.copy()
    score_cols = [c for c in full_df.columns if "Score" in c]
    full_df = full_df.drop(columns=score_cols, errors="ignore")
    
    # Rename columns: T1_ -> GSS, T2_ -> Opera
    rename_map_full = {}
    for col in full_df.columns:
        if col.startswith(PREFIX_T1):
            rename_map_full[col] = col.replace(PREFIX_T1, LABEL_GSS, 1)
        elif col.startswith(PREFIX_T2):
            rename_map_full[col] = col.replace(PREFIX_T2, LABEL_OPERA, 1)
    full_df = full_df.rename(columns=rename_map_full)
    
    for r in dataframe_to_rows(full_df, index=False, header=True):
        ws_full.append(r)
    
    # Format date columns in full data sheet too
    date_cols_full = []
    for idx, col_name in enumerate(full_df.columns, start=1):
        col_letter = get_column_letter(idx)
        if "date" in col_name.lower() or "arrival" in col_name.lower() or "departure" in col_name.lower():
            date_cols_full.append(col_letter)
            ws_full.column_dimensions[col_letter].width = 12
    
    for col_letter in date_cols_full:
        for row in ws_full.iter_rows(min_row=2, min_col=ord(col_letter) - ord('A') + 1, max_col=ord(col_letter) - ord('A') + 1):
            for cell in row:
                if cell.value is not None:
                    if isinstance(cell.value, str):
                        try:
                            for fmt in ["%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y", "%d.%m.%y"]:
                                try:
                                    dt = datetime.strptime(cell.value, fmt)
                                    cell.value = dt
                                    cell.number_format = numbers.FORMAT_DATE_DDMMYY
                                    break
                                except ValueError:
                                    continue
                            if isinstance(cell.value, str) and not isinstance(cell.value, datetime):
                                try:
                                    dt_val = pd.to_datetime(cell.value, dayfirst=False, errors='coerce')
                                    if pd.notna(dt_val):
                                        cell.value = dt_val.to_pydatetime() if hasattr(dt_val, 'to_pydatetime') else dt_val
                                        cell.number_format = numbers.FORMAT_DATE_DDMMYY
                                except:
                                    pass
                        except:
                            pass
                    elif isinstance(cell.value, (datetime, pd.Timestamp)):
                        if isinstance(cell.value, pd.Timestamp):
                            cell.value = cell.value.to_pydatetime()
                        cell.number_format = numbers.FORMAT_DATE_DDMMYY
                    elif isinstance(cell.value, (int, float)):
                        try:
                            if 1 <= cell.value <= 1000000:
                                dt = from_excel(cell.value)
                                cell.value = dt
                                cell.number_format = numbers.FORMAT_DATE_DDMMYY
                        except:
                            pass
                    elif hasattr(cell.value, 'date'):
                        cell.value = cell.value.date() if hasattr(cell.value, 'date') else cell.value
                        cell.number_format = numbers.FORMAT_DATE_DDMMYY
    
    # Auto-size other columns in full sheet
    for idx, col_name in enumerate(full_df.columns, start=1):
        col_letter = get_column_letter(idx)
        if col_letter not in date_cols_full:
            max_len = len(str(col_name))
            if len(full_df) > 0:
                sample_len = max([len(str(val)) for val in full_df[col_name].head(10).values if val is not None], default=0)
                max_len = max(max_len, sample_len)
            ws_full.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

    try:
        wb.save(path)
        log(f"[ExportCustomizer] Exported Excel to {path}")
    except Exception as e:
        log(f"[ExportCustomizer] Failed to save Excel: {e}")
